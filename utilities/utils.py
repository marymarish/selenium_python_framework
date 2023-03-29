import inspect
import logging
import softest
from openpyxl import Workbook, load_workbook
import csv

# anything which is independent of the driver reference
class Utils(softest.TestCase):
    global logger
    def loggen(loglevel=logging.INFO):
        # set class/method name from where its called:
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        # create console handler or file handler and set the log level:
        #ch = logging.StreamHandler()  # on console
        fh = logging.FileHandler(".\\Logs\\automations.log\\")  #file handler
        # create formatter
        formatter = logging.Formatter("%(asctime)s - %(levelname)s : %(message)s",
                            datefmt="%m/%d/%Y %I:%M:%S %p")
        # add formatter to console of file handler:
        #ch.setFormatter(formatter)
        fh.setFormatter(formatter)
        # add console handler to logger
        #logger.addHandler(ch)
        logger.addHandler(fh)
        return logger

    @staticmethod
    def read_from_excel(file_name, sheet):
        datalist = []
        work_book = load_workbook(filename=file_name)
        sheet = work_book[sheet]
        row_number = sheet.max_row
        column_number = sheet.max_column

        for i in range(2, row_number + 1):
            row = []
            for j in range(1, column_number + 1):
                row.append(sheet.cell(row=1, column=j).value)
            datalist.append(row)
        return datalist

    @staticmethod
    def read_data_from_csv(filename):
        # create emply list:
        data_list = []
        # open CSV file, "r" - read mode
        csvdata = open(filename, "r")
        # create CSV reader:
        reader = csv.reader(csvdata)
        # skip header:
        next(reader)
        # add CSV rows to empty list
        for row in reader:
            data_list.append(row)
        return data_list

